from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 打开现有的工作簿
wb = load_workbook("colored_rows.xlsx")
# 选择默认的工作表
ws = wb.active

# 定义两种不同的背景颜色
color1 = "DAEEF3"
color2 = "FDE9D9"
odd_fill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
even_fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')

# 遍历每一行，根据奇偶行设置不同的背景颜色
for row_idx, row in enumerate(ws.iter_rows(), start=1):
    if row_idx % 2 == 0:  # 偶数行
        for cell in row:
            cell.fill = even_fill
    else:  # 奇数行
        for cell in row:
            cell.fill = odd_fill

# 保存工作簿
wb.save("colored_rows.xlsx")
